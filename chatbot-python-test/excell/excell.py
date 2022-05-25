import openpyxl

wb = openpyxl.load_workbook('./sample.xlsx')
sheet = wb['Sheet1']
print(sheet.max_column, sheet.max_row)
print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=2, column=1).value)

# min_row : 2번째 행부터 탐색 시작
for row in sheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value)
    print('-' * 10)

# A2에서 C3 셀까지 슬라이싱 후 지정한 범위만큼 셀 데이터를 가져온다
cells = sheet['A2':'C3']
for row in cells:
    for cell in row:
        print(cell.value)
    print('-' * 10)

wb = openpyxl.Workbook()
sheet = wb.active
# 워크 시트 제목 지정
sheet.title = '회원정보'
# 표 헤더 컬럼 저장
header_titles = ['이름', '전화번호']
for idx, title in enumerate(header_titles):
        sheet.cell(row=1, column=idx+1, value=title)
# 표 내용 저장
members = [('kei', '010-1234-1234'), ('hong', '010-4321-1234')]
row_num = 2
for r, member in enumerate(members):
    for c, v in enumerate(member):
        sheet.cell(row=row_num+r, column=c+1, value=v)
wb.save('./members.xlsx')
wb.close()